"""
Shared pytest fixtures and synthetic data builders.

Test fixtures create synthetic files with fake (non-PHI) data that follow
the exact naming conventions and directory structures expected by the CCLF,
MSSP, MCQM, and BNEX glob patterns.
No real patient data is used anywhere in this test suite.
"""

import csv
import io
import zipfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import openpyxl

import duckdb
import pytest

from mssp_pipeline.processing.defs.cclf_file_defs import CCLFFileDef
from mssp_pipeline.processing.defs.mssp_file_defs import MSSPFileDef

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ACO_ID = "T0000"
TEST_DATE_STR = "250101"  # YYMMDD embedded in CCLF filenames → 2025-01-01
TEST_DATE = date(2025, 1, 1)


# ---------------------------------------------------------------------------
# Synthetic data builders
# ---------------------------------------------------------------------------


def build_cclf_line(file_def: CCLFFileDef, values: dict) -> str:
    """Build a fixed-width CCLF line from a dict of {column_name: value}.
    Columns not present in values are space-padded."""
    line = ""
    for col in file_def.columns:
        val = str(values.get(col.name, "")).ljust(col.width)[: col.width]
        line += val
    return line


def build_csv_bytes(headers: list, rows: list) -> bytes:
    """Encode a list of rows as CSV bytes with the given headers."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode()


# ---------------------------------------------------------------------------
# Session wrapper
# ---------------------------------------------------------------------------


class _TestSession:
    """Minimal DuckDBSession stand-in used in tests.

    Installs the zipfs extension (required to read zip archives) and sets the
    split character, mirroring what DuckDBSession does in production.
    """

    def __init__(self):
        self.connection = duckdb.connect(":memory:")
        self.connection.execute(
            "INSTALL zipfs FROM community; LOAD zipfs; SET zipfs_split = '!';"
        )
        self.connection.execute("INSTALL excel; LOAD excel;")
        self.connection.execute("INSTALL rusty_sheet FROM community; LOAD rusty_sheet;")

    def close(self):
        self.connection.close()


@pytest.fixture
def test_session():
    sess = _TestSession()
    yield sess
    sess.close()


# ---------------------------------------------------------------------------
# Config fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def raw_dir(tmp_path):
    return tmp_path


@pytest.fixture
def test_config(raw_dir):
    return SimpleNamespace(
        ACO_ID=ACO_ID,
        FILE_STORE=str(raw_dir),
        OUTPUT_TYPE="DUCKDB",
        FULL_REFRESH=True,
    )


@pytest.fixture
def incremental_config(raw_dir):
    return SimpleNamespace(
        ACO_ID=ACO_ID,
        FILE_STORE=str(raw_dir),
        OUTPUT_TYPE="DUCKDB",
        FULL_REFRESH=False,
    )


# ---------------------------------------------------------------------------
# CCLF plain-text file factory
# ---------------------------------------------------------------------------


def make_cclf_file(
    raw_dir: Path,
    file_def: CCLFFileDef,
    rows: list,
    date_str: str = TEST_DATE_STR,
    bundle_suffix: str = "",
) -> Path:
    """Create a CCLF plain-text fixture mirroring the new delivery structure:

        raw_dir/T0000/2025/P.T0000.ACO.ZCY25{bundle_suffix}/
            P.T0000.ACO.{pattern}.D{date_str}.T000000001  ← fixed-width plain text

    CCLF files are no longer zipped — they are plain fixed-width text files
    written directly to a year/bundle directory hierarchy. bundle_suffix
    varies the bundle directory name so incremental tests can create files
    with distinct FILE_PATHs (e.g. bundle_suffix="_first", "_second").

    The bundle directory name deliberately omits a .D<date>. segment to avoid
    confusing the FILE_DATE regex (which must match only the file's own date).

    Returns the path to the created file.
    """
    bundle_dir = raw_dir / ACO_ID / "2025" / "ZCY25" / f"P.{ACO_ID}.ACO.ZCY25{bundle_suffix}"
    bundle_dir.mkdir(parents=True, exist_ok=True)

    file_name = f"P.{ACO_ID}.ACO.{file_def.filename_pattern}.D{date_str}.T000000001"
    content = "\n".join(build_cclf_line(file_def, row) for row in rows).encode()

    file_path = bundle_dir / file_name
    file_path.write_bytes(content)
    return file_path


# ---------------------------------------------------------------------------
# MSSP zip factory
# ---------------------------------------------------------------------------


def _mssp_internal_name(pattern: str) -> str:
    """Convert a glob pattern like 'ALR*-1.csv' to 'ALR_test-1.csv'."""
    return pattern.replace("*", "_test")


# ---------------------------------------------------------------------------
# MCQM xlsx zip factory
# ---------------------------------------------------------------------------


def make_mcqm_zip(
    raw_dir: Path,
    sheet_data: dict,
    quarter: str = "2025Q4",
    date_str: str = "259999",
    time_str: str = "0400000",
) -> Path:
    """Create a MCQM zip fixture mirroring the new delivery structure:

        raw_dir/T0000/2025/P.T0000.QMCQM.D{date_str}.T{time_str}/
            P.T0000.ACO.MCQM.{quarter}.D{date_str}.T{time_str}  ← zip (no extension)
                └── P.T0000.ACO.MCQM.{quarter}.D{date_str}.T{time_str}.xlsx

    The zip file has no extension and sits inside a bundle directory, matching
    the actual CMS delivery layout. sheet_data is a dict of
    {sheet_name: (headers_list, rows_list)}.
    Returns the path to the created (extension-less) zip file.
    """
    base = f"P.{ACO_ID}.ACO.MCQM.{quarter}.D{date_str}.T{time_str}"
    xlsx_name = f"{base}.xlsx"
    bundle_dir_name = f"P.{ACO_ID}.QMCQM.D{date_str}.T{time_str}"
    bundle_dir = raw_dir / ACO_ID / "2025" / "QMCQM" / bundle_dir_name
    bundle_dir.mkdir(parents=True, exist_ok=True)

    # Build xlsx
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet
    for sheet_name, (headers, rows) in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    xlsx_bytes = io.BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    # Build single zip (no extension) directly containing the xlsx
    zip_path = bundle_dir / base
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr(xlsx_name, xlsx_bytes.read())
    return zip_path


# ---------------------------------------------------------------------------
# EXPU xlsx factory
# ---------------------------------------------------------------------------


def make_expu_xlsx(
    raw_dir: Path,
    sheet_data: dict,
    quarter: str = "2025Q1",
    date_str: str = "259999",
    time_str: str = "0100000",
) -> Path:
    """Create an EXPU xlsx fixture mirroring the delivery structure:

        raw_dir/T0000/2025/P.T0000.ACO.QEXPU.{quarter}.D{date_str}.T{time_str}/
            P.T0000.ACO.QEXPU.{quarter}.D{date_str}.T{time_str}.xlsx

    Unlike MCQM, the xlsx is written directly into the bundle directory — there
    is no additional zip wrapper. sheet_data is a dict of
    {sheet_name: [[row], [row], ...]} where rows are plain lists (no separate
    headers list — all rows including title rows are passed as data since the
    processor reads with header=false).

    Returns the path to the created xlsx file.
    """
    base = f"P.{ACO_ID}.ACO.QEXPU.{quarter}.D{date_str}.T{time_str}"
    bundle_dir = raw_dir / ACO_ID / "2025" / "QEXPU" / base
    bundle_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet
    for sheet_name, rows in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

    xlsx_path = bundle_dir / f"{base}.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def make_mssp_zip(
    raw_dir: Path,
    file_def: MSSPFileDef,
    headers: list,
    rows: list,
    bundle_suffix: str = "_Q1",
) -> Path:
    """Create the new three-level MSSP structure:

        raw_dir/2025/P.T0000.ACO.QEXPU<bundle_suffix>.D259999.T0000000/
            P.T0000.ACO.<TYPE>.D259999.T0000000   ← zip archive (no .zip extension)
                └── P.T0000.ACO.<TYPE>.D259999.T0000000.csv

    The zip filename uses no extension, matching the actual CMS delivery format.
    <TYPE> is derived from file_def.filename_pattern (e.g. 'BEUR*.csv' -> 'BEUR').

    Returns the path to the created zip.
    """
    type_prefix = file_def.filename_pattern.split("*")[0]
    bundle_dir_name = f"P.{ACO_ID}.ACO.QEXPU{bundle_suffix}.D259999.T0000000"
    bundle_dir = raw_dir / ACO_ID / "2025" / "QEXPU" / bundle_dir_name
    bundle_dir.mkdir(parents=True, exist_ok=True)

    # Embed bundle_suffix in the zip archive name so that multiple bundles for the
    # same report type produce distinct zip filenames (matching real CMS delivery
    # where each quarterly bundle carries the quarter in the zip filename).
    # The internal CSV name is derived from filename_pattern so the !*{pattern} glob
    # matches it correctly (e.g. 'ALR*-1.csv' -> 'ALR_test-1.csv').
    zip_name = f"P.{ACO_ID}.ACO.{type_prefix}{bundle_suffix}.D259999.T0000000"
    internal_name = _mssp_internal_name(file_def.filename_pattern)
    content = build_csv_bytes(headers, rows)

    zip_path = bundle_dir / zip_name
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.writestr(internal_name, content)
    return zip_path
